
from fastapi import Request, FastAPI, status, HTTPException, Depends
from fastapi.exceptions import RequestValidationError
from fastapi.responses import JSONResponse
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from models import Base, engine
from schemas import Role, AuthContext, GroupFormModel
from auth import auth_permission, auth_token, protect_api
from dotenv import load_dotenv
from datetime import datetime
import schemas
import models
import tasks
import uvicorn
import os
from stat_ import format_time
import openpyxl
from openpyxl.styles import Font, Alignment
import pandas as pd
from exports import export_pdf

app = FastAPI()

load_dotenv()
Base.metadata.create_all(bind=engine)
# create folder
RESULT_FOLDER = os.path.join(os.path.dirname(__file__), "result")
app.mount("/images", StaticFiles(directory=RESULT_FOLDER), name="result")

app = FastAPI()

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.exception_handler(RequestValidationError)
async def validation_exception_handler(request: Request, exc: RequestValidationError):
    detail = ""
    try:
        # detail = str(exc.errors())
        detail = exc.errors()[0]['loc'][-1]+','+exc.errors()[0]['msg']
    except Exception as e:
        print(e)
        detail = str(exc.errors())
    return JSONResponse({'detail': detail},
                        status_code=status.HTTP_422_UNPROCESSABLE_ENTITY)

# --------------------------------- Groups ---------------------------------#


@app.get("/groups", response_model=schemas.GroupListModel)
async def get_groups(page: int | None = 1, limit: int | None = 25, auth: AuthContext = Depends(auth_token)):
    if auth.role == Role.SUPER:
        return models.get_groups_no_org(page=page, size=limit)
    elif auth.role in (Role.AUDITOR, Role.ADMIN):
        return models.get_groups_by_org(auth.organization, page, limit)
    # read main database
    return models.get_groups_by_user(id=auth.id, page=page, size=limit)


@app.get("/groups/summary", response_model=schemas.GroupSumListModel)
async def get_group_sum(id: int, page: int | None = 1,
                        limit: int | None = 25, auth: AuthContext = Depends(auth_token)):
    if auth.role == Role.SUPER:
        return models.get_sum_group_no_org(page=page, size=limit)
    elif auth.role in (Role.AUDITOR, Role.ADMIN):
        return models.get_sum_groups_by_org(auth.organization, page, limit)
    # read main database
    return models.get_sum_groups_by_user(id=auth.id, page=page, size=limit)


@app.get("/groups/{id}", response_model=schemas.GroupDisplayModel)
async def get_group(id: int, auth: AuthContext = Depends(auth_token)):
    group = models.get_group_by_id(id)
    if not group:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Not Found group of id :{id}")
    if auth.role == Role.SUPER:
        return group
    elif auth.role in (Role.AUDITOR, Role.ADMIN):
        if group.org_id != None and group.org_id != auth.organization:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=f"Not Found group of id :{id}")
        return group
    elif group.org_id != None or (
            group.user_id != None and group.user_id != auth.id):
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Not Found group of id :{id}")
    return group


@app.get("/groups/{id}/summary", response_model=schemas.GroupSumModel)
async def get_group(id: int, auth: AuthContext = Depends(auth_token)):
    group = models.get_sum_groups_by_id(id)
    if not group:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Not Found group of id :{id}")
    if auth.role == Role.SUPER:
        return group
    elif auth.role in (Role.AUDITOR, Role.ADMIN):
        if group.org_id != None and group.org_id != auth.organization:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=f"Not Found group of id :{id}")
        return group
    elif group.org_id != None or (
            group.user_id != None and group.user_id != auth.id):
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Not Found group of id :{id}")
    return group


@app.post("/groups", response_model=schemas.GroupDisplayModel)
async def create_group(group: schemas.GroupModel, auth: AuthContext = Depends(auth_token)):
    group.modified_date = datetime.now()
    if auth.role == Role.SUPER:
        if not group.user_id:
            group.user_id = auth.id
        if not group.org_id:
            group.org_id = auth.organization
        group = models.create_group(group)
        if group:
            return group
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="invalid group format")
    group.user_id = auth.id
    group.org_id = None
    if auth.role == Role.ADMIN:
        group.org_id = auth.organization
    elif auth.role == Role.AUDITOR:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="UNAUTHORIZED")
    elif auth.role == Role.PAID:
        if models.count_groups_by_user(auth.id) >= 3:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="already at maximum number of Group")
        if len(group.targets) >= 50:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="targets can only be 50 per group")
    elif auth.role == Role.GUEST:
        if models.count_group_by_user(auth.id) >= 1:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="already at maximum number of Group")
        if len(group.targets) >= 10:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="targets can only be 10 per group")
    group.user_id = auth.id
    group.org_id = auth.organization
    group = models.create_group(group)
    if group:
        return group
    raise HTTPException(
        status_code=status.HTTP_400_BAD_REQUEST,
        detail="invalid group format")


@app.put("/groups/{id}", response_model=schemas.GroupDisplayModel)
async def modify_group(id: int, group: GroupFormModel, auth: AuthContext = Depends(auth_token)):
    auth_permission(auth, roles=(
        Role.ADMIN, Role.SUPER, Role.GUEST, Role.PAID))
    org_id = models.get_org_of_group(id)
    d = {}
    if group.name:
        d['name'] = group.name

    if auth.role == Role.SUPER:
        s = models.update_group(id, d, add_targets=group.targets_to_add,
                                remove_targets=group.targets_to_remove,
                                org_id=org_id, max_target=-1)
        if s:
            return s
    elif auth.role == Role.ADMIN:
        g = models.get_group_by_id(id)
        if g and g.org_id == auth.organization:
            s = models.update_group(id, d, add_targets=group.targets_to_add,
                                    remove_targets=group.targets_to_remove,
                                    org_id=auth.organization, max_target=-1)
            if s:
                return s
    elif auth.role in (Role.GUEST, Role.PAID):
        m = 10
        if auth.role == Role.PAID:
            m == 50
        g = models.get_group_by_id(id)
        if g and g.user_id == auth.id:
            s = models.update_group(id, d, add_targets=group.targets_to_add,
                                    remove_targets=group.targets_to_remove,
                                    org_id=auth.organization, max_target=m)
            if s:
                return s
    raise HTTPException(
        status_code=status.HTTP_404_NOT_FOUND,
        detail="Group not found")


@app.delete("/groups/{id}")
async def delete_smtp_config(id: int, auth: AuthContext = Depends(auth_token)):
    auth_permission(auth, roles=(
        Role.ADMIN, Role.SUPER, Role.GUEST, Role.PAID))
    org_id = models.get_org_of_group(id)
    if auth.role == Role.SUPER:
        if models.delete_group(id, org_id=org_id):
            return {'success': True}
    elif auth.role == Role.ADMIN:
        group = models.get_group_by_id(id)
        if group and group.org_id == auth.organization:
            if models.delete_group(id, auth.organization):
                return {'success': True}
    elif auth.role in (Role.GUEST, Role.PAID):
        group = models.get_group_by_id(id)
        if group and group.user_id == auth.id:
            if models.delete_group(id, org_id=None):
                return {'success': True}

    raise HTTPException(
        status_code=status.HTTP_404_NOT_FOUND,
        detail="Group not found")


# @app.post("/groups/import")


# --------------------------------  Campaigns --------------------------------#


@app.get("/campaigns", response_model=schemas.CampaignListModel)
async def get_campaigns(page: int | None = 1, limit: int | None = 25, auth: AuthContext = Depends(auth_token)):
    if auth.role == Role.SUPER:
        return models.get_all_campaigns(page, limit)
    elif auth.role in (Role.AUDITOR, Role.ADMIN):
        return models.get_campaigns_by_org(auth.organization, page, limit)
    # read main database
    return models.get_campaigns_by_user(id=auth.id, page=page, size=limit)


@app.get("/campaigns/summary", response_model=schemas.CampaignSumListModel)
def get_campaigns_sum(page: int | None = 1, limit: int | None = 25, auth: AuthContext = Depends(auth_token)):
    if auth.role == Role.SUPER:
        return models.get_all_campaigns_sum(page=page, size=limit)
    elif auth.role in (Role.AUDITOR, Role.ADMIN):
        return models.get_campaigns_sum_by_org(auth.organization, page=page, size=limit)
    return models.get_campaigns_sum_by_user(auth.id)


@app.get("/campaigns/{id}", response_model=schemas.CampaignDisplayModel)
async def get_campaign(id: int, auth: AuthContext = Depends(auth_token)):
    campaign = models.get_campaign_by_id(id)
    if not campaign:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Not Found campaign of id :{id}")
    if auth.role == Role.SUPER:
        return campaign
    elif auth.role in (Role.AUDITOR, Role.ADMIN):
        if campaign.org_id != None and campaign.org_id != auth.organization:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=f"Not Found campaign of id :{id}")
        return campaign
    elif campaign.org_id != None or (
            campaign.user_id != None and campaign.user_id != auth.id):
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Not Found campaign of id :{id}")
    return campaign


@app.post("/campaigns", response_model=schemas.CampaignDisplayModel)
async def create_campaign(campaign: schemas.CampaignModel, auth: AuthContext = Depends(auth_token)):
    campaign.create_date = datetime.now()
    if auth.role == Role.SUPER:
        if not campaign.user_id:
            campaign.user_id = auth.id
        if not campaign.org_id:
            campaign.org_id = auth.organization
        campaign = models.create_campaign(campaign)
        if campaign:
            return campaign
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="invalid campaign format")
    campaign.user_id = auth.id
    campaign.org_id = None
    if auth.role == Role.ADMIN:
        campaign.org_id = auth.organization
    elif auth.role == Role.AUDITOR:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="UNAUTHORIZED")
    elif auth.role == Role.PAID:
        if models.count_campaign_by_user(auth.id) >= 3:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="already at maximum number of Campaign")
    elif auth.role == Role.GUEST:
        if models.count_campaign_by_user(auth.id) >= 1:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="already at maximum number of Campaign")
    campaign.user_id = auth.id
    campaign.org_id = auth.organization
    campaign = models.create_campaign(campaign)
    if campaign:
        return campaign
    raise HTTPException(
        status_code=status.HTTP_400_BAD_REQUEST,
        detail="invalid campaign format")


@app.put("/campaigns/{id}", response_model=schemas.CampaignDisplayModel)
async def modify_campaign(id: int, campaign: schemas.CampaignFormModel, auth: AuthContext = Depends(auth_token)):
    auth_permission(auth, roles=(
        Role.ADMIN, Role.SUPER, Role.GUEST, Role.PAID))
    if auth.role == Role.SUPER:
        c = models.update_campaign(
            id, cam_in=campaign.dict(exclude_unset=True))
        if c:
            return c
    elif auth.role == Role.ADMIN:
        c = models.get_campaign_by_id(id)
        if c and c.org_id == auth.organization:
            c = models.update_campaign(
                id, cam_in=campaign.dict(exclude_unset=True))
            if c:
                return c
    elif auth.role == Role.GUEST or auth.role == Role.PAID:
        c = models.get_campaign_by_id(id)
        if c and c.user_id == auth.id:
            c = models.update_campaign(
                id, cam_in=campaign.dict(exclude_unset=True))
            if c:
                return c

    raise HTTPException(
        status_code=status.HTTP_404_NOT_FOUND,
        detail="SMTP not found")


@app.delete("/campaigns/{id}")
async def delete_campaign(id: int, auth: AuthContext = Depends(auth_token)):
    auth_permission(auth, roles=(
        Role.ADMIN, Role.SUPER, Role.GUEST, Role.PAID))
    if auth.role == Role.SUPER:
        campaign = models.delete_campaign(id)
        if campaign:
            return {'success': True}
    elif auth.role == Role.ADMIN:
        campaign = models.get_campaign_by_id(id)
        if campaign and campaign.org_id == auth.organization:
            campaign = models.delete_campaign(id)
            if campaign:
                return {'success': True}
    elif auth.role == Role.GUEST or auth.role == Role.PAID:
        campaign = models.get_campaign_by_id(id)
        if campaign and campaign.user_id == auth.id:
            campaign = models.delete_campaign(id)
            if campaign:
                return {'success': True}

    raise HTTPException(
        status_code=status.HTTP_404_NOT_FOUND,
        detail=f"Campaign not found :{id}")


@app.get("/campaigns/{id}/results", response_model=schemas.CampaignResultModel)
def get_campaign_result(id: int, auth: AuthContext = Depends(auth_token)):
    camp = models.get_campaign_by_id(id)
    if not camp:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Campaign :{id} not found")
    if auth.role == Role.SUPER:
        camp = models.get_campaign_result_by_id(id)
        if not camp:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=f"Campaign :{id} not found")
        return camp

    if auth.role in (Role.ADMIN, Role.AUDITOR):
        if camp.org_id != auth.organization:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=f"Campaign :{id} not found")
        camp = models.get_campaign_result_by_id(id)
        if camp:
            return camp
    elif auth.role in (Role.PAID, Role.GUEST):
        if camp.user_id != auth.id:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=f"Campaign :{id} not found")
        camp = models.get_campaign_result_by_id(id)
        if camp:
            return camp
    raise HTTPException(
        status_code=status.HTTP_404_NOT_FOUND,
        detail=f"Campaign :{id} not found")


@app.get("/campaigns/{id}/summary", response_model=schemas.CampaignSummaryModel)
def get_campaign_sum(id: int, auth: AuthContext = Depends(auth_token)):
    camp = models.get_campaign_sum_by_id(id)
    if not camp:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Campaign :{id} not found")
    if auth.role == Role.SUPER:
        return camp

    if auth.role in (Role.ADMIN, Role.AUDITOR):
        if camp.org_id != auth.organization:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Campaign :{id} not found")
        return camp
    elif auth.role in (Role.PAID, Role.GUEST):
        if camp.user_id != auth.id:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Campaign :{id} not found")
        return camp
    raise HTTPException(
        status_code=status.HTTP_404_NOT_FOUND,
        detail="Campaign :{id} not found")


@app.post("/campaigns/{id}/complete")
async def commplete_campaign(id: int, auth: AuthContext = Depends(auth_token)):
    camp = models.get_campaign_by_id(id)
    if not camp:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Campaign :{id} not found")
    if camp.status != schemas.Status.RUNING:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Campaign is alredy completed")
    if auth.role == Role.SUPER:
        camp.completed_date = datetime.now()
        await tasks.stop_campaign_tasks(campaign=camp, auth=auth)
        models.update_campaign(id, cam_in={'completed_date': camp.completed_date,
                                           'status': schemas.Status.COMPLETE})
        return {'success': True}
    if auth.role == Role.ADMIN:
        if camp.org_id != auth.organization:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Campaign :{id} not found")
    elif auth.role in (Role.PAID, Role.GUEST):
        if camp.user_id != auth.id:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Campaign :{id} not found")
    camp.completed_date = datetime.now()
    await tasks.stop_campaign_tasks(campaign=camp, auth=auth)
    models.update_campaign(id, cam_in={'completed_date': camp.completed_date,
                                       'status': schemas.Status.COMPLETE})
    return {'success': True}


@app.post("/campaigns/{id}/launch")
async def launch(id: int, auth: AuthContext = Depends(auth_token)):
    camp = models.get_campaign_by_id(id)
    if not camp:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Campaign :{id} not found")
    if camp.status != schemas.Status.IDLE:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Campaign is alredy running")
    groups = models.get_group_by_id(camp.group_id)

    if not groups or groups.org_id != camp.org_id:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Campaign has missinng data")
    if not groups.targets:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Campaign has no targets")
    if auth.role == Role.SUPER:
        result = await tasks.lanuch_campaign(camp, groups.targets, auth)
        models.add_results(data=result, camp=camp)
        models.update_campaign(id, cam_in={'launch_date': datetime.now(),
                                           'status': schemas.Status.RUNING})
        return {'success': True}
    if auth.role == Role.ADMIN:
        if camp.org_id != auth.organization:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Campaign :{id} not found")
    elif auth.role in (Role.PAID, Role.GUEST):
        if camp.user_id != auth.id:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Campaign :{id} not found")
    await tasks.lanuch_campaign(camp, groups.targets, auth)
    models.update_campaign(id, cam_in={'launch_date': datetime.now(),
                                       'status': schemas.Status.RUNING})
    return {'success': True}


@app.post("/event")
def event_callback(event: schemas.EventModel, _=Depends(protect_api)):
    print(event.dict())
    camp = models.get_campaign_by_id(event.campaign_id)
    if not camp or camp.status == schemas.Status.COMPLETE or event.email == None or event.r_id == None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Campaign not found")
    event.time = datetime.now()
    if models.update_result(org_id=camp.org_id, event=event):
        models.add_event(event=event)
        return {'success': True}
    raise HTTPException(
        status_code=status.HTTP_404_NOT_FOUND,
        detail="Target's result not found")


@app.put("/org")
def update_org():
    pass


@app.get("/campaigns/results/all")
def get_all_campaign_results(auth: AuthContext = Depends(auth_token)):
    if auth.role == Role.SUPER:
        all_results = models.get_all_result()
        return models.count_status(all_results)
    elif auth.role in (Role.AUDITOR, Role.ADMIN):
        org_results = models.get_result_by_id()
        return models.count_status(org_results)
    user_results = models.get_result_by_id()
    return models.count_status(user_results)

############### For Export to Excel ###############

# can't change for role


@app.get("/campaigns/{id}/results/export")
def get_results(id: int, auth: AuthContext = Depends(auth_token)):
    camp = models.get_campaign_sum_by_id(id)
    same = models.get_campaign_result_by_id_for_export(id)
    results, timelines, statistics = same

    overview_df = pd.DataFrame(camp, columns=['Events', 'Count'])

    results_df = pd.DataFrame(results)
    # drop data
    results_df = results_df.drop(['r_id', 'firstname', 'lastname', 'status',
                                 'ip', 'latitude', 'longitude', 'position', 'modified_date'], axis=1)
    results_df = results_df.fillna('FALSE')

    # Apply time formatting
    time_columns = ['time_sent_to_submit', 'time_sent_to_open',
                    'time_open_to_click', 'time_click_to_submit', 'time_sent_to_report']
    for col in time_columns:
        results_df[col] = results_df[col].apply(format_time)

    timelines_df = pd.DataFrame(timelines)
    timelines_df = timelines_df.drop(['campaign_id', 'r_id'], axis=1)
    statistics_df = pd.DataFrame(statistics.items(), columns=['List', 'Value'])
    statistics_df = statistics_df.fillna('FALSE')

    # Convert time columns to timedelta
    time_columns = [
        'mean_time_sent_to_submit', 'mean_time_sent_to_open', 'mean_time_open_to_click',
        'mean_time_click_to_submit', 'mean_time_sent_to_report',
        'std_time_sent_to_submit', 'std_time_sent_to_open', 'std_time_open_to_click',
        'std_time_click_to_submit', 'std_time_sent_to_report'
    ]

    for col in time_columns:
        statistics_df.loc[statistics_df['List'] == col, 'Value'] = format_time(
            statistics_df[statistics_df['List'] == col]['Value'].values[0])
    # Define the Excel export path
    # excel_path = r"C:\Users\Panda_X\Downloads\sheets_export_day_12.xlsx"

    # iname = ''.join(choices(ascii_letters, k=8))
    # results_filename = iname + '.xlsx'
    results_filename = "Phishing Results for Campaign.xlsx"
    excel_path = os.path.join(RESULT_FOLDER, results_filename)

    # Calculate the maximum length of data in each column
    max_lengths = results_df.apply(lambda col: col.astype(str).str.len().max())

    # Create Excel writer and export data to sheets
    with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
        overview_df.to_excel(writer, sheet_name='Overview', index=False)
        results_df.to_excel(writer, sheet_name='Results', index=False)
        timelines_df.to_excel(writer, sheet_name='Timelines', index=False)
        statistics_df.to_excel(writer, sheet_name='Statistics', index=False)

        # Access the workbook
        workbook = writer.book

        # Set header background color to blue
        header_color = openpyxl.styles.PatternFill(
            start_color="0066CC", end_color="0066CC", fill_type="solid")
        white_font = Font(color="FFFFFF")  # Define a white font

        # Apply header background color and white font color to each sheet
        for sheet in workbook.worksheets:
            for cell in sheet.iter_rows(min_row=1, max_row=1):
                for c in cell:
                    c.fill = header_color
                    c.font = white_font  # Apply the white font color to the cell text

            # Adjust column dimensions based on maximum data length
            for i, column in enumerate(results_df.columns, start=1):
                column_letter = openpyxl.utils.get_column_letter(i)
                column_width = max_lengths[column] * \
                    1.2  # Adjust the width as needed
                sheet.column_dimensions[column_letter].width = column_width
                # Center align data in cells
                # Skip the header row
                for row_cells in sheet.iter_rows(min_row=2):
                    for cell in row_cells:
                        cell.alignment = Alignment(
                            horizontal="center", vertical="center")
    return {'success': True}

# --------------------Export-PDF-------------------------#


# @app.get("/campaigns/{id}/result/export_pdf", responses={
#     200: {"content": {"application/pdf": {}}}
# })
@app.get("/campaigns/{id}/result/export_pdf")
def get_campaign_sum(id: int, auth: AuthContext = Depends(auth_token)):
    camp = models.get_campaign_by_id(id)
    if not camp:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Campaign :{id} not found")
    if auth.role == Role.SUPER:
        return export_pdf(camp)

    if auth.role in (Role.ADMIN, Role.AUDITOR):
        if camp.org_id != auth.organization:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=f"Campaign :{id} not found")
        return camp
    elif auth.role in (Role.PAID, Role.GUEST):
        if camp.user_id != auth.id:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=f"Campaign :{id} not found")
        return camp
    raise HTTPException(
        status_code=status.HTTP_404_NOT_FOUND,
        detail=f"Campaign :{id} not found")


if __name__ == "__main__":
    models.init_org_db()
    uvicorn.run(app, host=os.getenv('HOST'), port=int(os.getenv('PORT')))
